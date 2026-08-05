#!/usr/bin/env python3
# -*- coding: utf-8 -*-

import json
import os

import openpyxl


BASE_DIR = os.path.dirname(os.path.abspath(__file__))
XLSX_PATH = os.path.join(BASE_DIR, "字母图纸.xlsx")
NUM_XLSX_PATH = os.path.join(BASE_DIR, "数字图纸.xlsx")
OUT_PATH = os.path.join(BASE_DIR, "letters.js")

# (rmin, rmax, cmin, cmax, baseline_row)
# baseline_row is the Excel row where the letter's baseline sits.
LOWER_REGIONS = {
    "a": (3, 9, 4, 8, 9),
    "b": (3, 9, 12, 16, 9),
    "c": (3, 9, 19, 23, 9),
    "d": (3, 9, 26, 30, 9),
    "e": (3, 9, 34, 38, 9),
    "f": (3, 9, 41, 45, 9),
    "g": (3, 12, 49, 53, 9),
    "h": (3, 9, 57, 61, 9),
    "i": (3, 9, 65, 65, 9),
    "j": (3, 10, 68, 70, 9),
    "k": (3, 9, 74, 77, 9),
    "l": (3, 9, 81, 81, 9),
    "m": (14, 18, 2, 8, 18),
    "n": (14, 18, 12, 16, 18),
    "o": (14, 18, 19, 23, 18),
    "p": (14, 21, 26, 30, 18),
    "q": (14, 21, 33, 37, 18),
    "r": (14, 18, 41, 44, 18),
    "s": (14, 18, 49, 53, 18),
    "t": (14, 18, 56, 60, 18),
    "u": (14, 18, 63, 67, 18),
    "v": (14, 18, 69, 73, 18),
    "w": (22, 25, 4, 8, 25),
    "x": (22, 26, 12, 16, 26),
    "y": (22, 29, 19, 23, 26),
    "z": (22, 26, 28, 32, 26),
}

UPPER_REGIONS = {
    "A": (33, 39, 4, 8, 39),
    "B": (33, 39, 12, 16, 39),
    "C": (33, 39, 20, 24, 39),
    "D": (33, 39, 29, 33, 39),
    "E": (33, 39, 37, 41, 39),
    "F": (33, 39, 45, 49, 39),
    "G": (33, 39, 53, 57, 39),
    "H": (33, 39, 61, 65, 39),
    "I": (33, 39, 69, 73, 39),
    "J": (33, 39, 77, 81, 39),
    "K": (33, 39, 85, 89, 39),
    "L": (33, 39, 93, 97, 39),
    "M": (43, 49, 2, 8, 49),
    "N": (43, 49, 12, 16, 49),
    "O": (43, 49, 20, 24, 49),
    "P": (43, 49, 29, 33, 49),
    "Q": (43, 49, 37, 41, 49),
    "R": (43, 49, 45, 49, 49),
    "S": (43, 49, 52, 56, 49),
    "T": (43, 49, 59, 63, 49),
    "U": (43, 49, 66, 70, 49),
    "V": (43, 49, 74, 78, 49),
    "W": (43, 49, 81, 85, 49),
    "X": (43, 49, 88, 92, 49),
    "Y": (53, 59, 4, 8, 59),
    "Z": (53, 59, 11, 15, 59),
}

# (rmin, rmax, cmin, cmax, baseline_row)
DIGIT_REGIONS = {
    "1": (5, 11, 3, 7, 11),
    "2": (5, 11, 9, 13, 11),
    "3": (5, 11, 17, 21, 11),
    "4": (5, 11, 24, 28, 11),
    "5": (5, 11, 31, 35, 11),
    "6": (15, 21, 3, 7, 21),
    "7": (15, 21, 9, 13, 21),
    "8": (15, 21, 17, 21, 21),
    "9": (15, 21, 24, 28, 21),
    "0": (15, 21, 31, 35, 21),
}

# Manual corrections supplied by the user. These win over the workbook extraction.
PATTERN_OVERRIDES = {
    "w": {
        "baseline": 4,
        "pattern": [
            [1, 0, 1, 0, 1],
            [1, 0, 1, 0, 1],
            [1, 0, 1, 0, 1],
            [1, 0, 1, 0, 1],
            [0, 1, 0, 1, 0],
        ],
    },
    "t": {
        "baseline": 5,
        "pattern": [
            [0, 0, 1, 0, 0],
            [0, 0, 1, 0, 0],
            [1, 1, 1, 1, 1],
            [0, 0, 1, 0, 0],
            [0, 0, 1, 0, 0],
            [0, 0, 1, 1, 1],
        ],
    },
    "Q": {
        "baseline": 6,
        "pattern": [
            [0, 1, 1, 1, 0, 0],
            [1, 0, 0, 0, 1, 0],
            [1, 0, 0, 0, 1, 0],
            [1, 0, 0, 0, 1, 0],
            [1, 0, 0, 1, 1, 0],
            [1, 0, 0, 0, 1, 0],
            [0, 1, 1, 1, 0, 1],
        ],
    },
    "L": {
        "baseline": 6,
        "pattern": [
            [1, 0, 0, 0],
            [1, 0, 0, 0],
            [1, 0, 0, 0],
            [1, 0, 0, 0],
            [1, 0, 0, 0],
            [1, 0, 0, 0],
            [1, 1, 1, 1],
        ],
    },
    "W": {
        "baseline": 6,
        "pattern": [
            [1, 0, 0, 0, 0, 0, 1],
            [1, 0, 0, 1, 0, 0, 1],
            [1, 0, 0, 1, 0, 0, 1],
            [1, 0, 0, 1, 0, 0, 1],
            [1, 0, 0, 1, 0, 0, 1],
            [1, 0, 0, 1, 0, 0, 1],
            [0, 1, 1, 0, 1, 1, 0],
        ],
    },
    "y": {
        "baseline": 4,
        "pattern": [
            [1, 0, 0, 0, 1],
            [1, 0, 0, 0, 1],
            [1, 0, 0, 0, 1],
            [1, 0, 0, 1, 1],
            [0, 1, 1, 0, 1],
            [0, 0, 0, 0, 1],
            [1, 0, 0, 0, 1],
            [0, 1, 1, 1, 0],
        ],
    },
}


def extract_letter(ws, region):
    rmin, rmax, cmin, cmax, baseline = region
    cells = {
        (r, c)
        for r in range(rmin, rmax + 1)
        for c in range(cmin, cmax + 1)
        if ws.cell(row=r, column=c).fill.fill_type not in (None, "none")
    }
    if not cells:
        raise ValueError("empty region")

    rows = [r for r, _ in cells]
    cols = [c for _, c in cells]
    rrmin, rrmax = min(rows), max(rows)
    ccmin, ccmax = min(cols), max(cols)

    pattern = []
    for r in range(rrmin, rrmax + 1):
        pattern.append([
            1 if (r, c) in cells else 0
            for c in range(ccmin, ccmax + 1)
        ])

    return {
        "pattern": pattern,
        "width": len(pattern[0]),
        "height": len(pattern),
        "baseline": baseline - rrmin,
    }


def main():
    wb = openpyxl.load_workbook(XLSX_PATH, data_only=True)
    ws = wb["Sheet1"]
    num_wb = openpyxl.load_workbook(NUM_XLSX_PATH, data_only=True)
    num_ws = num_wb["Sheet1"]

    letters = {}
    for key, region in list(LOWER_REGIONS.items()) + list(UPPER_REGIONS.items()):
        letters[key] = extract_letter(ws, region)
    for key, region in DIGIT_REGIONS.items():
        letters[key] = extract_letter(num_ws, region)
    for key, override in PATTERN_OVERRIDES.items():
        pattern = override["pattern"]
        letters[key] = {
            "pattern": pattern,
            "width": len(pattern[0]),
            "height": len(pattern),
            "baseline": override["baseline"],
        }

    styles = {
        "classic": letters,
    }
    payload = {
        "letters": letters,
        "styles": styles,
        "generated_from": os.path.basename(XLSX_PATH),
    }
    with open(OUT_PATH, "w", encoding="utf-8") as fh:
        fh.write("// Generated by build_letters.py. Do not edit by hand.\n")
        fh.write("window.LETTERS = ")
        fh.write(json.dumps(payload, ensure_ascii=False, separators=(",", ":")))
        fh.write(";\n")

    print("wrote", OUT_PATH, "letters:", len(letters))


if __name__ == "__main__":
    main()
